# -*- coding: utf-8 -*-
"""
Created on Thu Jul 22 11:58:09 2021

@author: kyle.conrad
"""

from Shipment import Shipment
from Helpers import Helpers
from pathlib import Path

import time, re, shutil, subprocess, traceback, sys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from datetime import date, datetime


class RusalShipment(Shipment):
    
    testingDest = "C://Users//kyleb//Documents//Clerical Applications//Testing Destination"
    testingReleasesFold = "C://Users//kyleb//Documents//Clerical Applications//Testing Releases"
    
    def __init__(self, driver, driverName, truckingCompany, dor, licenseNumber, clientCode, loadNum):
        self.driver = driver
        self.driverName = driverName
        self.truckingCompany = truckingCompany
        self.dor = dor
        self.licenseNumber= licenseNumber
        self.clientCode = clientCode
        self.loadNum = loadNum
    
    def set_product(self):
        try:
            print("\nSetting product type...")
            if "INGOTS" in self.convertedFile:
                self.product = "INGOTS"
            else:
                self.product = "BILLETS"
            print("Product type successfully set as " + self.product)
        except:
            Helpers.Exceptions.unexpected_exception("set product type")
            
    def set_poNumber(self):
        try:
            print("\nSetting PO Number...")
            if self.product == "INGOTS":
                self.poNumber = re.findall(r'\d*\s/\sS-\d*\s/\s\w*\'\d*', self.convertedFile)[0]
            else: 
                try:
                    self.poNumber = re.findall(r'\d*-\d{1,2}\"[x,X]\d*\"/\d\w{2}/\w{3}/\d*-\d*-\d*\s/\sS-\d*\s/\s\w*\'\d*', self.convertedFile)[0]
                except IndexError:
                    self.poNumber = re.findall(r'\d*\s/\sS-\d*\s/\s\w{3}\'\d{2}', self.convertedFile)[0]
            print("PO Number set as " + self.poNumber)
        except:
            Helpers.Exceptions.unexpected_exception("set PO number")
            
    def set_remarks(self):
        try:
            print("\nSetting remarks...")
            self.set_clerk()
            self.set_poNumber()
            self.remarks = "{}\nPO#: {}\nMaterial must be free of dirt debris.\nMetal straps must be replaced with plastic straps.\nDrivers accept paperwork as is, signing waived due to COVID-19 restrictions.\n{}".format(self.dor, self.poNumber, self.clerk)
            print("\nRemarks successfully set")
        except:
            Helpers.Exceptions.unexpected_exception("set remarks")
            
    def set_pickupNumber(self):
        self.pickupNumber = self.dor
        
    #incomplete needs methods for dealing with multiple piece counts and for dealing with billets
    def set_shippedBundles(self):
        try:
            print("\nSetting bundle count...")
            if self.product == "INGOTS":
                self.shippedBundles = re.findall(r'\d{1,2} BUNDLES', self.convertedFile)[0].replace(" BUNDLES", "")
                
            elif self.product == "BILLETS": 
                self.convertedFile = re.sub(r'PC', 'PIECE', re.sub(r'TK','TRUCK', self.convertedFile))
                self.shippedBundles = re.findall(r'\d{1,2}\sTRUCKS\s@\s\d{1,2}\sPIECES', self.convertedFile)
                if (len(self.shippedBundles) > 0):
                    #code to determine if short load or normal load, use load number supplied by user at start
                    "Something"
                else:
                    self.shippedBundles = re.findall(r'\d{1,2}\sPIECES/TRUCK', self.convertedFile)[0].replace(" PIECES/TRUCK", "")
            print("Successfully set bundle count as " + self.shippedBundles)
        except:
            print("An unexpected exception occcurred while attempting to set bundle amount: \n")
            print(traceback.print_exc())
                
                
    def create_scale_ticket(self):
        try:
            print("\nCreating scale ticket...")
            file = self.dirPath + "/Resources/Rusal Billet Scale Template.xlsx"
            
            #destination = self.testingDest + "\\Scale Ticket " + self.driverName + ".xlsx"
            
            destination = "Z:\\SCALE OFFICE\\RUSAL\\Current Billet Scale Tickets\\Scale Ticket " + self.driverName + ".xlsx"
            
            shutil.copyfile(file, destination)
            
            workbook = load_workbook(filename = destination)
            sheet = workbook.active
            
            sheet.cell(21, 3).value = date.today().strftime("%m/%d/%Y")
            sheet.cell(23, 3).value = self.driverName
            sheet.cell(26, 3).value = self.dor
            sheet.cell(27, 3).value = self.poNumber
            sheet.cell(21, 9).value = datetime.now().strftime("%H:%M %p")
            sheet.cell(23, 9).value = self.truckingCompany
            sheet.cell(24, 9).value = self.licenseNumber
            
            workbook.save(destination)
            workbook.close()
            print("Scale ticket successfully created for " + self.driverName)
            
        except PermissionError:
            print("Could not create scale ticket as scale ticket with same name is already open by another user!")
            
        except:
            print("An unexpected exception occurred while adding creating scale ticket: \n")
            print(traceback.print_exc())
        
    def set_variables(self):
        #self.convert_to_text(self.testingReleasesFold + "\\" + self.dor + ".pdf")
        
        self.convert_to_text("Z:/SCALE OFFICE/RUSAL/Current Releases/" + self.dor + ".pdf")
        self.set_product()
        self.set_pickupNumber()
        self.set_remarks()
        self.set_carrierBill()
        self.set_address()
        self.set_receiver()
        self.set_shippedBundles()
        if self.product == "BILLETS":
            self.create_scale_ticket()
    
    def createEmail(self):
        try:
            print("\nAttempting to create email draft...")
            path = "C:\\Users\\kyle.conrad\\Documents\\Kyle Conrad\\Python Scripts\\Rusal Email Draft.ps1"
            Helpers.Email.main(path, "-order", self.dor, "-load", self.loadNum)
            print("Email draft successfully created for " + self.loadNum)
            
        except:
            Helpers.Exceptions.unexpected_exception("create email draft")
        
    #incomplete needs methods for dealing with multiple B/Ls on site and for dealing with multiple piece counts
    def add_load(self):
        
        self.createEmail()
        
        try:
            print("\nAttempting to add load...")
            WebDriverWait(self.driver,50).until(lambda x: x.find_element(By.XPATH,'//*[@id="viewport"]/article/section/form/section[1]/div[3]/nav/menu/a[2]'))
            self.driver.get('https://tos.qsl.com/client-inventories/shipment-of-materials/' + self.bolNum + '/shipped-items')
            element = WebDriverWait(self.driver, 50).until(lambda x: x.find_element(By.XPATH, '//*[@id="viewport"]/article/section/article/header/aside/a[2]'))
            element.click()
            element = WebDriverWait(self.driver, 50).until(lambda x: x.find_element(By.XPATH, '/html/body/div[18]/div/div/article/form/div/div/button'))
            element.click()
            element = WebDriverWait(self.driver, 50).until(lambda x: x.find_element(By.XPATH, '/html/body/div[18]/div/div/article/section/div/div/div[1]/div[4]/div/div[2]/span/button'))
            time.sleep(2.5)
            element.click()
            element = WebDriverWait(self.driver, 10).until(lambda x: x.find_element(By.XPATH, '//*[@id="quantity"]'))
            element.send_keys(self.shippedBundles  + Keys().RETURN)
            time.sleep(.5)
            self.driver.find_element_by_xpath('/html/body/div[18]/div/div/div[1]/header/menu/button[3]').click()
            time.sleep(2)
            self.driver.get('https://tos.qsl.com/client-inventories/shipment-of-materials/' + self.bolNum + '/shipped-items')
            print("\nLoad successfully added")
            
        except KeyboardInterrupt:
            Helpers.Exceptions.user_interrupt(self.driver)
    
        except:
            Helpers.Exceptions.unexpected_exception("add load")
        