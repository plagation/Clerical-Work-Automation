# -*- coding: utf-8 -*-
"""
Created on Fri Jul 23 07:28:11 2021

@author: kyle.conrad
"""

import Shipment
import time
from pathlib import Path
import re
from openpyxl import load_workbook
from datetime import date
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

class BoscusTruckReception(Shipment.Shipment):
    dirPath = str(Path.cwd()).replace("\\", "/")
    
    def set_remarks(self):
        self.set_clerk()
        self.remarks = "{}\nOffloaded on {}\n{}".format(self.dor, date.today().strftime("%m/%d/%Y"), self.clerk)
    
    def parse_pdf(self):
        
        #Specific to network (Modification for scale necessary)
        self.convert_to_text("Z:/SCALE OFFICE/Boscus/Current Releases/" + self.dor + "-VTE-B80-000.pdf")
        self.descriptions = re.findall(r'\d{1}x\d{1}x\d+[\'( 92\-5/8)]+[\'\- 96]+', self.convertedFile)
        i = 0
        while i < len(self.descriptions):
            self.descriptions[i] = self.descriptions[i].replace("'", "").replace(" ", "")
            if "92" in self.descriptions[i]:
                self.descriptions[i] = self.descriptions[i][0:4] + "92 5/8"
            elif "96" in self.descriptions[i]:
                self.descriptions[i] = self.descriptions[i][0:4] + "96"
            i += 1
        
        self.pieceCounts = re.findall(r'\s{2}\d{3}\s{2}', self.convertedFile)

        self.bundleCounts = re.findall(r'\s{2}\d{1,2}\s{2}', self.convertedFile)
        
        if len(self.bundleCounts) > len(self.pieceCounts):
            self.bundleCounts.pop(len(self.bundleCounts)-1)
            
        i = 0 
        while i < len(self.bundleCounts):
            self.bundleCounts[i] = self.bundleCounts[i].replace(" ", "")
            self.pieceCounts[i] = self.pieceCounts[i].replace(" ", "")
            i += 1
            
        self.volumes = [None] * len(self.descriptions)
        self.weights = [None] * len(self.descriptions)
        i = 0
        while i < len(self.descriptions):
            sizeArr = self.descriptions[i].split("x")
            self.volumes[i] = 1
            for ft in sizeArr:
                if ("96" in ft) or ("92" in ft):
                    self.volumes[i] = self.volumes[i] * 8
                else:
                    self.volumes[i] = self.volumes[i] * int(ft)
            self.volumes[i] = (self.volumes[i] * int(self.pieceCounts[i]))/12
            self.weights[i] = self.volumes[i] * 1.75
            i += 1
                
    def form_manifest(self):
        self.parse_pdf()
        workbook = load_workbook(filename = self.dirPath + "/import-manifest.xlsx")
        sheet = workbook.active
        
        while sheet.cell(3,1).value != None:
            sheet.delete_rows(3)
            
        i = 0
        while i < len(self.pieceCounts):
            row = 3 + i
            sheet.cell(row,1).value = "Boscus"
            sheet.cell(row,2).value = "Nasco Stock"
            sheet.cell(row,3).value = "Lumber"
            sheet.cell(row,4).value = "Dimensional Lumber"
            sheet.cell(row,5).value = "Unit"
            sheet.cell(row,6).value = self.pieceCounts[i]
            sheet.cell(row,7).value = "Can be containerized"
            sheet.cell(row,9).value = self.bundleCounts[i]
            sheet.cell(row,10).value = "Description"
            sheet.cell(row,11).value = self.descriptions[i]
            sheet.cell(row,12).value = "Other"
            sheet.cell(row,13).value = date.today().strftime("%m/%d/%Y")
            sheet.cell(row,30).value = self.volumes[i]
            sheet.cell(row,31).value = "FBM"
            sheet.cell(row,32).value = self.weights[i]
            sheet.cell(row,33).value = "lb"
            sheet.cell(row,34).value = "East Pavement"
            i += 1
        
        workbook.save(self.dirPath + "/import-manifest.xlsx")
        workbook.close()
            
    def CreateShipment(self):
        self.set_remarks()
        self.form_manifest()
        self.driver.get("https://tos.qsl.com/client-inventories/receptions-of-materials")
        WebDriverWait(self.driver,50).until(lambda x: x.find_element(By.XPATH, '//*[@id="viewport"]/article/section/section[1]/div[2]/div[2]/button'))
        time.sleep(1.5)
        self.driver.find_element_by_xpath('//*[@id="viewport"]/article/section/section[1]/div[2]/div[2]/button').click()
        WebDriverWait(self.driver, 10).until(lambda x: x.find_element(By.XPATH, '/html/body/div[3]/div/form/header/menu/button[2]'))
        self.driver.find_element_by_xpath('/html/body/div[3]/div/form/header/menu/button[2]').click()
        WebDriverWait(self.driver, 50).until(lambda x: x.find_element(By.XPATH, '//*[@id="react-select-5-input"]'))
        self.driver.find_element_by_xpath('//*[@id="react-select-5-input"]').send_keys(self.truckingCompany + Keys().RETURN)
        self.driver.find_element_by_xpath('//*[@id="driverName"]').send_keys(self.driverName)
        self.driver.find_element_by_xpath('//*[@id="carrierBill"]').send_keys(self.dor)
        self.driver.find_element_by_xpath('//*[@id="transportationNumber"]').send_keys(self.licenseNumber)
        self.driver.find_element_by_xpath('//*[@id="specialInstructions"]').send_keys(self.remarks)
        self.driver.find_element_by_xpath('//*[@id="react-select-13-input"]').send_keys("Boscus" + Keys().RETURN)
        self.driver.find_element_by_xpath('//*[@id="viewport"]/article/section/form/section[2]/div/div[2]/button[2]').click()
        WebDriverWait(self.driver, 50).until(lambda x: x.find_element(By.XPATH, '//*[@id="viewport"]/article/section/form/section[2]/div/div[1]/h3'))
        self.bolNum = self.driver.find_element_by_xpath('//*[@id="viewport"]/article/section/form/section[2]/div/div[1]/h3').text
        self.driver.get('https://tos.qsl.com/client-inventories/receptions-of-materials/' + self.bolNum + '/incoming-items')
        WebDriverWait(self.driver, 50).until(lambda x: x.find_element(By.XPATH, '//*[@id="viewport"]/article/section/section[3]/div[2]/div[4]/button'))
        time.sleep(1.5)
        self.driver.find_element_by_xpath('//*[@id="viewport"]/article/section/section[3]/div[2]/div[1]').click()
        self.driver.find_element(By.CSS_SELECTOR, 'input[type="file"]').send_keys(self.dirPath + "/import-manifest.xlsx")